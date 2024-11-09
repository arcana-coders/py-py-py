import openpyxl

def extract_concepts(file_path):
    # Abre el archivo de Excel
    workbook = openpyxl.load_workbook(file_path)
    
    # Verifica si existe la hoja "Información"
    if "Información" not in workbook.sheetnames:
        print("La hoja 'Información' no existe.")
        return
    
    # Hoja donde se almacenarán los conceptos
    info_sheet = workbook["Información"]
    
    # Conjunto para almacenar conceptos únicos
    concepts_set = set()
    
    # Evitar conceptos relacionados con "mano de obra"
    excluded_terms = ["mano de obra", "mano de Obra", "Mano de obra", "MANO DE OBRA"]
    
    # Itera sobre todas las hojas, excepto la hoja de información
    for sheet in workbook.sheetnames:
        if sheet == "Información":
            continue
        
        ws = workbook[sheet]
        print(f"Procesando hoja: {sheet}")
        
        # Buscar los encabezados de DESCRIPCIÓN y PRECIO en las primeras 20 filas
        desc_col = None
        price_col = None
        header_row = None
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
            for idx, cell in enumerate(row):
                if isinstance(cell, str):
                    if cell.lower() == "descripción":
                        desc_col = idx
                    elif cell.lower() == "precio":
                        price_col = idx
            
            if desc_col is not None and price_col is not None:
                header_row = row_idx
                break
        
        if desc_col is not None and price_col is not None:
            print(f"Encontradas columnas 'DESCRIPCIÓN' en {desc_col} y 'PRECIO' en {price_col} en la fila {header_row}")
            # Itera sobre las filas de datos desde la fila donde se encontraron los encabezados
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                description = row[desc_col]
                price = row[price_col]
                
                if description and price:
                    if isinstance(description, str) and all(term.lower() not in description.lower() for term in excluded_terms):
                        # Agrega la descripción y el precio al conjunto de conceptos únicos
                        concepts_set.add((description, price))
                        print(f"Agregado concepto: {description} con precio {price}")
        else:
            print(f"No se encontraron las columnas 'DESCRIPCIÓN' y 'PRECIO' en la hoja: {sheet}")
    
    # Escribir los conceptos únicos en la hoja de información
    for idx, (description, price) in enumerate(sorted(concepts_set), start=1):
        info_sheet.cell(row=idx, column=1, value=description)
        info_sheet.cell(row=idx, column=2, value=price)
    
    # Guarda el archivo de Excel
    workbook.save(file_path)
    print("Conceptos únicos con precios escritos en la hoja 'Información'.")


# Ruta del archivo de Excel
file_path = "C:/Users/Capalsa/Desktop/pythondesktop/tallergus.xlsx"

extract_concepts(file_path)
