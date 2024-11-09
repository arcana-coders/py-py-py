import openpyxl
import re

def rename_sheet_by_content(file_path):
    # Abre el archivo de Excel
    workbook = openpyxl.load_workbook(file_path)
    
    # Itera sobre todas las hojas
    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        
        eco_number = None
        budget_number = None
        
        # Itera sobre todas las celdas en la hoja
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell:
                    # Busca el número económico
                    eco_match = re.search(r'ECO: ?(\d+)', str(cell))
                    if eco_match:
                        eco_number = eco_match.group(1)
                    
                    # Busca el número de presupuesto
                    budget_match = re.search(r'Presupuesto Nº (\d+)', str(cell))
                    if budget_match:
                        budget_number = budget_match.group(1)
            
            # Si ambos números se encuentran, no hace falta seguir buscando en esta hoja
            if eco_number and budget_number:
                break
        
        # Cambia el nombre de la hoja según los números encontrados
        if eco_number and budget_number:
            new_sheet_name = f"{eco_number}-{budget_number}"
        elif budget_number:
            new_sheet_name = budget_number
        else:
            continue  # No cambia el nombre de la hoja si no se encuentra ningún número
        
        # Cambia el nombre de la hoja y maneja posibles errores de longitud o caracteres no permitidos
        try:
            workbook[sheet].title = new_sheet_name
        except ValueError:
            # Si el nuevo nombre es muy largo, trúncalo
            workbook[sheet].title = new_sheet_name[:31]
    
    # Guarda el archivo de Excel con los nuevos nombres de las hojas
    workbook.save(file_path)

# Ruta del archivo de Excel
file_path = "C:/Users/Capalsa/Desktop/pythondesktop/tallergus.xlsx"

rename_sheet_by_content(file_path)

