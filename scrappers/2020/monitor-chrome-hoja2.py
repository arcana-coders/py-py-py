import openpyxl
import requests
from bs4 import BeautifulSoup as soup
# from colorama import Fore, init


def precioydisponibilidad():
    # se extrae el precio y se le da formato, se quita la coma, y las decimales y el simbolo
    price_1 = scrap.find(id='priceblock_ourprice')
    if price_1 is None:
        price_1 = scrap.find(id='priceblock_saleprice')
    if price_1 is None:
        price_1 = scrap.find(id='priceblock_dealprice')
    if price_1 is None:
        price_1 = scrap.find(id='newBuyBoxPrice')
    if price_1 is None:
        ws.cell(row, 9).value = pausada
        return
    print(price_1)
    price_4 = price_1.get_text()
    verted_price = price_4[1:10]
    onverted_price = verted_price.replace(',', '')
    sep = '.'
    converted_price = onverted_price.split(sep, 1)[0]
    ws.cell(row, 6).value = int(converted_price)
    print('precio proveedor: ' + str(converted_price))

    # arreglamos el precio con la formula
    rmula = (int(converted_price) * float(1.85))
    ormula = str(rmula)
    formula = ormula.split(sep, 1)[0]
    ws.cell(row, 7).value = int(formula)
    print('precio de venta: ' + formula)

    # Ahora la disponibilidad cuando es nacional o importado.
    disponibilidad = scrap.find(id='priceblock_ourprice_ifdmsg')
    if disponibilidad != None:
        disponibilidad = 5
    if disponibilidad == None:
        disponibilidad = 0
    ws.cell(row, 11).value = disponibilidad
    ws.cell(row, 10).value = 3
    print('Disponible en: ' + str(disponibilidad) + ' dias.')

# Primero las variables.

wb = openpyxl.load_workbook("cuentaarturo.xlsx")
ws = wb['Hoja2']

# Comenzamos el loop donde lee las columnas desde la celda 2 hasta donde tenga valor

for row in range(2,ws.max_row):
    if(ws.cell(row,8).value is None):
        break

    Asin = (ws.cell(row, 8).value)
    estado = (ws.cell(row, 9).value)  # sus opciones son 'Pausada' o 'Activa'
    pre_pro = (ws.cell(row, 6).value)
    precio = (ws.cell(row, 7).value)
    estate = {"1": "Activa", "2": "Pausada"}
    activa = 'Activa'
    pausada = 'Pausada'

    URL = 'https://www.amazon.com.mx/dp/' + Asin
    headers = {
        "User-Agent": 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.97 Safari/537.36'}

    print('')
    print(URL)
    print('Estado anterior: ' + estado)
    page = requests.get(URL, headers=headers)
    scrap = soup(page.content, 'html.parser')
    disponible = scrap.find(id='availability')
    enviado = scrap.find(id='merchant-info')
    # print(enviado)
    # init(autoreset=True)
    dispo = 'Activa'
    nodispo = 'Pausada'

    if disponible == None:
        print('No tiene informacion.')
    elif disponible != None:
        disponible = disponible.get_text().split()
        if disponible[0] == ('No') or enviado == None:
            ws.cell(row, 9).value = pausada
            print(nodispo)
        elif disponible[0] == ('Disponible'):
            ws.cell(row, 9).value = pausada
            print(nodispo)
        elif disponible[0] == ('Disponible.') and enviado is not None:
            enviado = enviado.get_text().split()
            if  any(ele == 'Amazon' for ele in enviado) or any(ele == 'Amazon.' for ele in enviado):
                ws.cell(row, 9).value = activa
                enviado = " ".join(enviado)
                print(dispo)
                print(enviado)
                precioydisponibilidad()
            else:
                ws.cell(row, 9).value = pausada
                enviado = " ".join(enviado)
                print(nodispo)
                print(enviado)
            disponible = " ".join(disponible)
            print(disponible)

wb.save('cuentaarturo.xlsx')