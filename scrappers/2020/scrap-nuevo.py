import requests
from bs4 import BeautifulSoup as soup
from datetime import datetime
import re
from colorama import init, Fore
import openpyxl

init(autoreset=True)

portada = Fore.LIGHTRED_EX + """
 ██████╗ █████╗ ██████╗  █████╗ ██╗    ███████╗ █████╗ 
██╔════╝██╔══██╗██╔══██╗██╔══██╗██║    ██╔════╝██╔══██╗
██║     ███████║██████╔╝███████║██║    ███████╗███████║
██║     ██╔══██║██╔═══╝ ██╔══██║██║       ══██╗██╔══██║
╚██████╗██║  ██║██║     ██║  ██║╚█████╗╚██████╔██║  ██║
 ╚═════╝╚═╝  ╚═╝╚═╝     ╚═╝  ╚═╝╚═════╝ ╚═════╝╚═╝  ╚═╝
                                                 Coders.
"""

saludo = Fore.GREEN + ('Bienvenidos a mi primer web scrapper')

print(portada)
print(saludo)

# El loop infinito del programa
while True:

    print('')
    print('')
    # nos pregunta sobre la busqueda y el link
    busqueda = input('Que busqueda realizaste? ')
    URL = input("inserta tu link: ")
    cabezal = {
        "User-Agent": 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.97 Safari/537.36'}

    # se abre la coneccion y se usa el modulo para extraer la info de la pagina
    page = requests.get(URL, headers=cabezal)
    page_soup = soup(page.content, 'html.parser')

    # agarra cada producto de la pagina de busqueda
    container1 = page_soup.findAll("div", {"data-component-type": "s-search-result"})

    # crea el archivo con la fecha y la busqueda
    fecha = str(datetime.now())
    fecha1 = fecha.replace(".", "_")
    fecha2 = fecha1.replace(' ', '_')
    fecha3 = fecha2.replace(':', '_')
    filename = busqueda.replace(' ', '_') + fecha3 + '.txt'
    f = open(filename, "w")

    # extrae el asin de cada container, cada producto en la pagina de busqueda.
    for contain in container1:
        amazin = contain["data-asin"]

        print("asin: " + amazin)

        f.write(amazin + ",")

    f.close()

    # la condicion para iniciar todo el programa nuevamente.
    print('')
    print('')
    reinicio = input('Quieres scrappear esa busqueda?  (si/no) ')
    if reinicio != 'si':
        continue
    if reinicio != 'no':
        break
    print('')
    print('')
    print('Ok, entonces hagamoslo de nuevo. ')

print('')
print('')
print('Entonces vamos a scrappear.')

#comenzamos a escrappear cada asin.

for contain in container1:
    amazin = contain["data-asin"]

    URL = 'https://www.amazon.com.mx/dp/' + amazin
    headers = {
        "User-Agent": 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.97 Safari/537.36'}

    print('')
    print(URL)
    page = requests.get(URL, headers=headers)
    scrap = soup(page.content, 'html.parser')
    disponible = scrap.find(id='availability')
    if disponible is not None:
        disponible = disponible.get_text().split()
    sindesc = scrap.find('ul', {'class': 'a-unordered-list a-vertical a-spacing-mini'})
    enviado = scrap.find(id='merchant-info')
    if enviado is None or any(ele != 'Disponible.' for ele in disponible):
        print("no esta disponible")
        print(disponible)
    elif enviado is not None:
        disponible = " ".join(disponible)
        enviado = enviado.get_text().split()
        print(disponible)
        # se extrae el precio y se le da formato, se quita la coma, y las decimales y el simbolo
        price_1 = scrap.find(id='priceblock_ourprice')
        if price_1 is None:
            price_1 = scrap.find(id='priceblock_saleprice')
        if price_1 is None:
            price_1 = scrap.find(id='priceblock_dealprice')
        if price_1 is None:
            price_1 = scrap.find(id='newBuyBoxPrice')
        if price_1 is None:
            print("no tiene precio disponible.")
        if price_1 is not None:
            price_4 = price_1.get_text()
            verted_price = price_4[1:10]
            onverted_price = verted_price.replace(',', '')
            sep = '.'
            converted_price = onverted_price.split(sep, 1)[0]
            if any(ele == 'Amazon' for ele in enviado) or any(ele == 'Amazon.' for ele in enviado):
                enviado = " ".join(enviado)
                if 100 <= int(converted_price) <= 999:
                    # las que no cambian en el excel
                    categoria = ('Otras Categorías > Otros')
                    estado = ('Activa')
                    stock = ('3')
                    tipo = ('Premium')
                    condicion = ('Nuevo')
                    envio = ('Sí')
                    modo = ('No Especificado')
                    metodo = ('Express a domicilio')
                    retiro = ('No')
                    garantia = ('30 dias de Garantia')
                    marca = ('Capalsa')

                    # se extrae el titulo del producto
                    titles = scrap.find(id="productTitle").get_text().strip()
                    titles = re.sub('[^a-zA-Z0-9 \n\.\:\´\¿\!\-\?\_\¡\%\ó\ñ\á\í\/\Á\Ó\é\(\)\\\ú]', '', titles)
                    titlex = titles.replace('Amazon', '')
                    title = titlex[0:59]

                    # arreglamos el precio con la formula
                    rmula = (int(converted_price) * float(1.85))
                    ormula = str(rmula)
                    formula = ormula.split(sep, 1)[0]

                    # Ahora la disponibilidad cuando es nacional o importado.
                    disponibilidad = scrap.find(id='priceblock_ourprice_ifdmsg')
                    if disponibilidad != None:
                        disponibilidad = 5
                    if disponibilidad == None:
                        disponibilidad = 0

                    # vamos a extraer la descripcion.
                    desc_1 = scrap.find('ul', {'class': 'a-unordered-list a-vertical a-spacing-mini'})
                    if desc_1 is not None:
                        desc_1 = desc_1.get_text().split()
                        # print(desc_1)
                    # la segunda descripcion.
                    desc_2 = scrap.find(id='productDescription')
                    if desc_2 is not None:
                        desc_2 = desc_2.get_text().split()

                    # Ahora unimos las descripciones. quitamos tablas y las limpiamos
                    desc_3 = (str(desc_1) + ' ' + str(desc_2))
                    desc_c = re.sub('[^a-zA-Z0-9 \n\.\:\´\¿\!\-\?\_\¡\%\ó\ñ\á\í\/\Á\Ó\é\(\)\\\ú]', '', desc_3)
                    desc_c = desc_c.replace("amazon", 'capalsa')
                    desc_c = desc_c.replace("Amazon", 'capalsa')
                    desc_c = desc_c.replace("amazon.", 'capalsa.')
                    desc_c = desc_c.replace("Amazon.", 'capalsa.')
                    desc_c = desc_c.replace("AMAZON", 'capalsa.')
                    desc_c = desc_c.replace("http", '')
                    desc_c = desc_c.replace("HTTP", '')
                    desc_c = desc_c.replace(".com", '')
                    desc_c = desc_c.replace(".COM", '')
                    desc_c = desc_c.replace("phone", '')
                    desc_d = desc_c.replace("@", '')

                    mensaje = ('COMERCIALIZADORA CAPALSA A SUS ORDENES ' + '\n'
                     'Todos nuestros productos son nuevos y originales. Garantia directa de 30 dias con nosotros ' + '\n'
                     'Facturamos si asi lo requiere. ' + '\n'
                     'Tiempos de entrega: ' + '\n'
                     '2 a 9 dias en productos nacionales (por contingencia) ' + '\n'
                     '7 a 15 dias en productos de importacion. ' + '\n')

                    titles = re.sub('[^a-zA-Z0-9 \n\.\:\´\¿\!\-\?\_\¡\%\ó\ñ\á\í\/\Á\Ó\é\(\)\\\ú]', '', titles)
                    titles = titles.replace("amazon", 'capalsa')
                    titles = titles.replace("Amazon", 'capalsa')
                    titles = titles.replace("amazon.", 'capalsa.')
                    titles = titles.replace("Amazon.", 'capalsa.')
                    titles = titles.replace("AMAZON", 'capalsa.')
                    titles = titles.replace("http", '')
                    titles = titles.replace(".com", '')
                    titles = titles.replace("phone", '')
                    titles = titles.replace("@", '')

                    desc_4 = mensaje + '\n' + titles + '\n' + '\n' + desc_d
                    # print('este articulo es para venderse en MELI.')
                    print(title)
                    print(enviado)
                    print('precio proveedor: ' + str(converted_price))
                    print('Precio MeLi: ' + str(formula))
                    print('Disponible en: ' + str(disponibilidad) + ' dias.')
                    # print(desc_4)

                    # Abre y escribe en las celdas y lo salva.
                    wb = openpyxl.load_workbook("test.xlsx")
                    ws = wb['Hoja1']
                    newRowLocation = ws.max_row + 1

                    # Extrae las imagenes

                    lista = scrap.select('.a-button-text img')
                    print('contiene: ' + str(len(lista)) + ' imagenes')
                    if (len(lista)) == 1:
                        lista_0 = lista[0]
                        lista_01 = (lista_0['src'])
                        lista_02 = lista_01.replace('_US40_', '_SX679_')
                        lista_02 = lista_02.replace('_SX38_SY50_CR,0,0,38,50_','_AC_SX679_')
                        lista_02 = lista_02.replace('_AC_SR38,50_', '_AC_SX679_')
                        ws.cell(column=(24), row=newRowLocation, value=lista_02)
                        print(lista_02)
                    if (len(lista)) == 2:
                        lista_0 = lista[0]
                        lista_01 = (lista_0['src'])
                        lista_02 = lista_01.replace('_US40_', '_SX679_')
                        lista_02 = lista_02.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                        lista_02 = lista_02.replace('_AC_SR38,50_', '_AC_SX679_')
                        ws.cell(column=(24), row=newRowLocation, value=lista_02)
                        print(lista_02)
                        if lista[1] != None:
                            lista_1 = lista[1]
                            lista_11 = (lista_1['src'])
                            lista_12 = lista_11.replace('_US40_', '_SX679_')
                            lista_12 = lista_12.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                            lista_12 = lista_12.replace('_AC_SR38,50_', '_AC_SX679_')
                            ws.cell(column=(25), row=newRowLocation, value=lista_12)
                            print(lista_12)
                    if (len(lista)) == 3:
                        lista_0 = lista[0]
                        lista_01 = (lista_0['src'])
                        lista_02 = lista_01.replace('_US40_', '_SX679_')
                        lista_02 = lista_02.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                        lista_02 = lista_02.replace('_AC_SR38,50_', '_AC_SX679_')
                        ws.cell(column=(24), row=newRowLocation, value=lista_02)
                        print(lista_02)
                        if lista[1] != None:
                            lista_1 = lista[1]
                            lista_11 = (lista_1['src'])
                            lista_12 = lista_11.replace('_US40_', '_SX679_')
                            lista_12 = lista_12.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                            lista_12 = lista_12.replace('_AC_SR38,50_', '_AC_SX679_')
                            ws.cell(column=(25), row=newRowLocation, value=lista_12)
                            print(lista_12)
                            if lista[2] != None:
                                lista_2 = lista[2]
                                lista_21 = (lista_2['src'])
                                lista_22 = lista_21.replace('_US40_', '_SX679_')
                                lista_22 = lista_22.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                                lista_22 = lista_22.replace('_AC_SR38,50_', '_AC_SX679_')
                                ws.cell(column=(26), row=newRowLocation, value=lista_22)
                                print(lista_22)
                    if (len(lista)) == 4:
                        lista_0 = lista[0]
                        lista_01 = (lista_0['src'])
                        lista_02 = lista_01.replace('_US40_', '_SX679_')
                        lista_02 = lista_02.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                        lista_02 = lista_02.replace('_AC_SR38,50_', '_AC_SX679_')
                        ws.cell(column=(24), row=newRowLocation, value=lista_02)
                        print(lista_02)
                        if lista[1] != None:
                            lista_1 = lista[1]
                            lista_11 = (lista_1['src'])
                            lista_12 = lista_11.replace('_US40_', '_SX679_')
                            lista_12 = lista_12.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                            lista_12 = lista_12.replace('_AC_SR38,50_', '_AC_SX679_')
                            ws.cell(column=(25), row=newRowLocation, value=lista_12)
                            print(lista_12)
                            if lista[2] != None:
                                lista_2 = lista[2]
                                lista_21 = (lista_2['src'])
                                lista_22 = lista_21.replace('_US40_', '_SX679_')
                                lista_22 = lista_22.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                                lista_22 = lista_22.replace('_AC_SR38,50_', '_AC_SX679_')
                                ws.cell(column=(26), row=newRowLocation, value=lista_22)
                                print(lista_22)
                                if lista[3] != None:
                                    lista_3 = lista[3]
                                    lista_31 = (lista_3['src'])
                                    lista_32 = lista_31.replace('_US40_', '_SX679_')
                                    lista_32 = lista_32.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                                    lista_32 = lista_32.replace('_AC_SR38,50_', '_AC_SX679_')
                                    ws.cell(column=(27), row=newRowLocation, value=lista_32)
                                    print(lista_32)
                    if (len(lista)) == 5:
                        lista_0 = lista[0]
                        lista_01 = (lista_0['src'])
                        lista_02 = lista_01.replace('_US40_', '_SX679_')
                        lista_02 = lista_02.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                        lista_02 = lista_02.replace('_AC_SR38,50_', '_AC_SX679_')
                        ws.cell(column=(24), row=newRowLocation, value=lista_02)
                        print(lista_02)
                        if lista[1] != None:
                            lista_1 = lista[1]
                            lista_11 = (lista_1['src'])
                            lista_12 = lista_11.replace('_US40_', '_SX679_')
                            lista_12 = lista_12.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                            lista_12 = lista_12.replace('_AC_SR38,50_', '_AC_SX679_')
                            ws.cell(column=(25), row=newRowLocation, value=lista_12)
                            print(lista_12)
                            if lista[2] != None:
                                lista_2 = lista[2]
                                lista_21 = (lista_2['src'])
                                lista_22 = lista_21.replace('_US40_', '_SX679_')
                                lista_22 = lista_22.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                                lista_22 = lista_22.replace('_AC_SR38,50_', '_AC_SX679_')
                                ws.cell(column=(26), row=newRowLocation, value=lista_22)
                                print(lista_22)
                                if lista[3] != None:
                                    lista_3 = lista[3]
                                    lista_31 = (lista_3['src'])
                                    lista_32 = lista_31.replace('_US40_', '_SX679_')
                                    lista_32 = lista_32.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                                    lista_32 = lista_32.replace('_AC_SR38,50_', '_AC_SX679_')
                                    ws.cell(column=(27), row=newRowLocation, value=lista_32)
                                    print(lista_32)
                                    if lista[4] != None:
                                        lista_4 = lista[4]
                                        lista_41 = (lista_4['src'])
                                        lista_42 = lista_41.replace('_US40_', '_SX679_')
                                        lista_42 = lista_42.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                                        lista_42 = lista_42.replace('_AC_SR38,50_', '_AC_SX679_')
                                        ws.cell(column=(28), row=newRowLocation, value=lista_42)
                                        print(lista_42)
                    if (len(lista)) == 6:
                        lista_0 = lista[0]
                        lista_01 = (lista_0['src'])
                        lista_02 = lista_01.replace('_US40_', '_SX679_')
                        lista_02 = lista_02.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                        lista_02 = lista_02.replace('_AC_SR38,50_', '_AC_SX679_')
                        ws.cell(column=(24), row=newRowLocation, value=lista_02)
                        print(lista_02)
                        if lista[1] != None:
                            lista_1 = lista[1]
                            lista_11 = (lista_1['src'])
                            lista_12 = lista_11.replace('_US40_', '_SX679_')
                            lista_12 = lista_12.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                            lista_12 = lista_12.replace('_AC_SR38,50_', '_AC_SX679_')
                            ws.cell(column=(25), row=newRowLocation, value=lista_12)
                            print(lista_12)
                            if lista[2] != None:
                                lista_2 = lista[2]
                                lista_21 = (lista_2['src'])
                                lista_22 = lista_21.replace('_US40_', '_SX679_')
                                lista_22 = lista_22.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                                lista_22 = lista_22.replace('_AC_SR38,50_', '_AC_SX679_')
                                ws.cell(column=(26), row=newRowLocation, value=lista_22)
                                print(lista_22)
                                if lista[3] != None:
                                    lista_3 = lista[3]
                                    lista_31 = (lista_3['src'])
                                    lista_32 = lista_31.replace('_US40_', '_SX679_')
                                    lista_32 = lista_32.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                                    lista_32 = lista_32.replace('_AC_SR38,50_', '_AC_SX679_')
                                    ws.cell(column=(27), row=newRowLocation, value=lista_32)
                                    print(lista_32)
                                    if lista[4] != None:
                                        lista_4 = lista[4]
                                        lista_41 = (lista_4['src'])
                                        lista_42 = lista_41.replace('_US40_', '_SX679_')
                                        lista_42 = lista_42.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                                        lista_42 = lista_42.replace('_AC_SR38,50_', '_AC_SX679_')
                                        ws.cell(column=(28), row=newRowLocation, value=lista_42)
                                        print(lista_42)
                                        if lista[5] != None:
                                            lista_5 = lista[5]
                                            lista_51 = (lista_5['src'])
                                            lista_52 = lista_51.replace('_US40_', '_SX679_')
                                            lista_52 = lista_52.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                                            lista_52 = lista_52.replace('_AC_SR38,50_', '_AC_SX679_')
                                            ws.cell(column=(29), row=newRowLocation, value=lista_52)
                                            print(lista_52)
                    if (len(lista)) == 7:
                        lista_0 = lista[0]
                        lista_01 = (lista_0['src'])
                        lista_02 = lista_01.replace('_US40_', '_SX679_')
                        lista_02 = lista_02.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                        lista_02 = lista_02.replace('_AC_SR38,50_', '_AC_SX679_')
                        ws.cell(column=(24), row=newRowLocation, value=lista_02)
                        print(lista_02)
                        if lista[1] != None:
                            lista_1 = lista[1]
                            lista_11 = (lista_1['src'])
                            lista_12 = lista_11.replace('_US40_', '_SX679_')
                            lista_12 = lista_12.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                            lista_12 = lista_12.replace('_AC_SR38,50_', '_AC_SX679_')
                            ws.cell(column=(25), row=newRowLocation, value=lista_12)
                            print(lista_12)
                            if lista[2] != None:
                                lista_2 = lista[2]
                                lista_21 = (lista_2['src'])
                                lista_22 = lista_21.replace('_US40_', '_SX679_')
                                lista_22 = lista_22.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                                lista_22 = lista_22.replace('_AC_SR38,50_', '_AC_SX679_')
                                ws.cell(column=(26), row=newRowLocation, value=lista_22)
                                print(lista_22)
                                if lista[3] != None:
                                    lista_3 = lista[3]
                                    lista_31 = (lista_3['src'])
                                    lista_32 = lista_31.replace('_US40_', '_SX679_')
                                    lista_32 = lista_32.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                                    lista_32 = lista_32.replace('_AC_SR38,50_', '_AC_SX679_')
                                    ws.cell(column=(27), row=newRowLocation, value=lista_32)
                                    print(lista_32)
                                    if lista[4] != None:
                                        lista_4 = lista[4]
                                        lista_41 = (lista_4['src'])
                                        lista_42 = lista_41.replace('_US40_', '_SX679_')
                                        lista_42 = lista_42.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                                        lista_42 = lista_42.replace('_AC_SR38,50_', '_AC_SX679_')
                                        ws.cell(column=(28), row=newRowLocation, value=lista_42)
                                        print(lista_42)
                                        if lista[5] != None:
                                            lista_5 = lista[5]
                                            lista_51 = (lista_5['src'])
                                            lista_52 = lista_51.replace('_US40_', '_SX679_')
                                            lista_52 = lista_52.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                                            lista_52 = lista_52.replace('_AC_SR38,50_', '_AC_SX679_')
                                            ws.cell(column=(29), row=newRowLocation, value=lista_52)
                                            print(lista_52)
                                            if lista[6] != None:
                                                lista_6 = lista[6]
                                                lista_61 = (lista_6['src'])
                                                lista_62 = lista_61.replace('_US40_', '_SX679_')
                                                lista_62 = lista_62.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                                                lista_62 = lista_62.replace('_AC_SR38,50_', '_AC_SX679_')
                                                ws.cell(column=(30), row=newRowLocation, value=lista_62)
                                                print(lista_62)
                    if (len(lista)) == 8:
                        lista_0 = lista[0]
                        lista_01 = (lista_0['src'])
                        lista_02 = lista_01.replace('_US40_', '_SX679_')
                        lista_02 = lista_02.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                        lista_02 = lista_02.replace('_AC_SR38,50_', '_AC_SX679_')
                        ws.cell(column=(24), row=newRowLocation, value=lista_02)
                        print(lista_02)
                        if lista[1] != None:
                            lista_1 = lista[1]
                            lista_11 = (lista_1['src'])
                            lista_12 = lista_11.replace('_US40_', '_SX679_')
                            lista_12 = lista_12.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                            lista_12 = lista_12.replace('_AC_SR38,50_', '_AC_SX679_')
                            ws.cell(column=(25), row=newRowLocation, value=lista_12)
                            print(lista_12)
                            if lista[2] != None:
                                lista_2 = lista[2]
                                lista_21 = (lista_2['src'])
                                lista_22 = lista_21.replace('_US40_', '_SX679_')
                                lista_22 = lista_22.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                                lista_22 = lista_22.replace('_AC_SR38,50_', '_AC_SX679_')
                                ws.cell(column=(26), row=newRowLocation, value=lista_22)
                                print(lista_22)
                                if lista[3] != None:
                                    lista_3 = lista[3]
                                    lista_31 = (lista_3['src'])
                                    lista_32 = lista_31.replace('_US40_', '_SX679_')
                                    lista_32 = lista_32.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                                    lista_32 = lista_32.replace('_AC_SR38,50_', '_AC_SX679_')
                                    ws.cell(column=(27), row=newRowLocation, value=lista_32)
                                    print(lista_32)
                                    if lista[4] != None:
                                        lista_4 = lista[4]
                                        lista_41 = (lista_4['src'])
                                        lista_42 = lista_41.replace('_US40_', '_SX679_')
                                        lista_42 = lista_42.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                                        lista_42 = lista_42.replace('_AC_SR38,50_', '_AC_SX679_')
                                        ws.cell(column=(28), row=newRowLocation, value=lista_42)
                                        print(lista_42)
                                        if lista[5] != None:
                                            lista_5 = lista[5]
                                            lista_51 = (lista_5['src'])
                                            lista_52 = lista_51.replace('_US40_', '_SX679_')
                                            lista_52 = lista_52.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                                            lista_52 = lista_52.replace('_AC_SR38,50_', '_AC_SX679_')
                                            ws.cell(column=(29), row=newRowLocation, value=lista_52)
                                            print(lista_52)
                                            if lista[6] != None:
                                                lista_6 = lista[6]
                                                lista_61 = (lista_6['src'])
                                                lista_62 = lista_61.replace('_US40_', '_SX679_')
                                                lista_62 = lista_62.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                                                lista_62 = lista_62.replace('_AC_SR38,50_', '_AC_SX679_')
                                                ws.cell(column=(30), row=newRowLocation, value=lista_62)
                                                print(lista_62)
                                                if lista[7] != None:
                                                    lista_7 = lista[7]
                                                    lista_71 = (lista_7['src'])
                                                    lista_72 = lista_71.replace('_US40_', '_SX679_')
                                                    lista_72 = lista_72.replace('_SX38_SY50_CR,0,0,38,50_',
                                                                                '_AC_SX679_')
                                                    lista_72 = lista_72.replace('_AC_SR38,50_', '_AC_SX679_')
                                                    ws.cell(column=(31), row=newRowLocation, value=lista_72)
                                                    print(lista_72)
                    if (len(lista)) == 9:
                        lista_0 = lista[0]
                        lista_01 = (lista_0['src'])
                        lista_02 = lista_01.replace('_US40_', '_SX679_')
                        lista_02 = lista_02.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                        lista_02 = lista_02.replace('_AC_SR38,50_', '_AC_SX679_')
                        ws.cell(column=(24), row=newRowLocation, value=lista_02)
                        print(lista_02)
                        if lista[1] != None:
                            lista_1 = lista[1]
                            lista_11 = (lista_1['src'])
                            lista_12 = lista_11.replace('_US40_', '_SX679_')
                            lista_12 = lista_12.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                            lista_12 = lista_12.replace('_AC_SR38,50_', '_AC_SX679_')
                            ws.cell(column=(25), row=newRowLocation, value=lista_12)
                            print(lista_12)
                            if lista[2] != None:
                                lista_2 = lista[2]
                                lista_21 = (lista_2['src'])
                                lista_22 = lista_21.replace('_US40_', '_SX679_')
                                lista_22 = lista_22.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                                lista_22 = lista_22.replace('_AC_SR38,50_', '_AC_SX679_')
                                ws.cell(column=(26), row=newRowLocation, value=lista_22)
                                print(lista_22)
                                if lista[3] != None:
                                    lista_3 = lista[3]
                                    lista_31 = (lista_3['src'])
                                    lista_32 = lista_31.replace('_US40_', '_SX679_')
                                    lista_32 = lista_32.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                                    lista_32 = lista_32.replace('_AC_SR38,50_', '_AC_SX679_')
                                    ws.cell(column=(27), row=newRowLocation, value=lista_32)
                                    print(lista_32)
                                    if lista[4] != None:
                                        lista_4 = lista[4]
                                        lista_41 = (lista_4['src'])
                                        lista_42 = lista_41.replace('_US40_', '_SX679_')
                                        lista_42 = lista_42.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                                        lista_42 = lista_42.replace('_AC_SR38,50_', '_AC_SX679_')
                                        ws.cell(column=(28), row=newRowLocation, value=lista_42)
                                        print(lista_42)
                                        if lista[5] != None:
                                            lista_5 = lista[5]
                                            lista_51 = (lista_5['src'])
                                            lista_52 = lista_51.replace('_US40_', '_SX679_')
                                            lista_52 = lista_52.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                                            lista_52 = lista_52.replace('_AC_SR38,50_', '_AC_SX679_')
                                            ws.cell(column=(29), row=newRowLocation, value=lista_52)
                                            print(lista_52)
                                            if lista[6] != None:
                                                lista_6 = lista[6]
                                                lista_61 = (lista_6['src'])
                                                lista_62 = lista_61.replace('_US40_', '_SX679_')
                                                lista_62 = lista_62.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                                                lista_62 = lista_62.replace('_AC_SR38,50_', '_AC_SX679_')
                                                ws.cell(column=(30), row=newRowLocation, value=lista_62)
                                                print(lista_62)
                                                if lista[7] != None:
                                                    lista_7 = lista[7]
                                                    lista_71 = (lista_7['src'])
                                                    lista_72 = lista_71.replace('_US40_', '_SX679_')
                                                    lista_72 = lista_72.replace('_SX38_SY50_CR,0,0,38,50_',
                                                                                '_AC_SX679_')
                                                    lista_72 = lista_72.replace('_AC_SR38,50_', '_AC_SX679_')
                                                    ws.cell(column=(31), row=newRowLocation, value=lista_72)
                                                    print(lista_72)
                                                    if lista[8] != None:
                                                        lista_8 = lista[8]
                                                        lista_81 = (lista_8['src'])
                                                        lista_82 = lista_81.replace('_US40_', '_SX679_')
                                                        lista_82 = lista_82.replace('_SX38_SY50_CR,0,0,38,50_',
                                                                                    '_AC_SX679_')
                                                        lista_82 = lista_82.replace('_AC_SR38,50_', '_AC_SX679_')
                                                        ws.cell(column=(32), row=newRowLocation, value=lista_82)
                                                        print(lista_82)
                    if (len(lista)) >= 10:
                        lista_0 = lista[0]
                        lista_01 = (lista_0['src'])
                        lista_02 = lista_01.replace('_US40_', '_SX679_')
                        lista_02 = lista_02.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                        lista_02 = lista_02.replace('_AC_SR38,50_', '_AC_SX679_')
                        ws.cell(column=(24), row=newRowLocation, value=lista_02)
                        print(lista_02)
                        if lista[1] != None:
                            lista_1 = lista[1]
                            lista_11 = (lista_1['src'])
                            lista_12 = lista_11.replace('_US40_', '_SX679_')
                            lista_12 = lista_12.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                            lista_12 = lista_12.replace('_AC_SR38,50_', '_AC_SX679_')
                            ws.cell(column=(25), row=newRowLocation, value=lista_12)
                            print(lista_12)
                            if lista[2] != None:
                                lista_2 = lista[2]
                                lista_21 = (lista_2['src'])
                                lista_22 = lista_21.replace('_US40_', '_SX679_')
                                lista_22 = lista_22.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                                lista_22 = lista_22.replace('_AC_SR38,50_', '_AC_SX679_')
                                ws.cell(column=(26), row=newRowLocation, value=lista_22)
                                print(lista_22)
                                if lista[3] != None:
                                    lista_3 = lista[3]
                                    lista_31 = (lista_3['src'])
                                    lista_32 = lista_31.replace('_US40_', '_SX679_')
                                    lista_32 = lista_32.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                                    lista_32 = lista_32.replace('_AC_SR38,50_', '_AC_SX679_')
                                    ws.cell(column=(27), row=newRowLocation, value=lista_32)
                                    print(lista_32)
                                    if lista[4] != None:
                                        lista_4 = lista[4]
                                        lista_41 = (lista_4['src'])
                                        lista_42 = lista_41.replace('_US40_', '_SX679_')
                                        lista_42 = lista_42.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                                        lista_42 = lista_42.replace('_AC_SR38,50_', '_AC_SX679_')
                                        ws.cell(column=(28), row=newRowLocation, value=lista_42)
                                        print(lista_42)
                                        if lista[5] != None:
                                            lista_5 = lista[5]
                                            lista_51 = (lista_5['src'])
                                            lista_52 = lista_51.replace('_US40_', '_SX679_')
                                            lista_52 = lista_52.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                                            lista_52 = lista_52.replace('_AC_SR38,50_', '_AC_SX679_')
                                            ws.cell(column=(29), row=newRowLocation, value=lista_52)
                                            print(lista_52)
                                            if lista[6] != None:
                                                lista_6 = lista[6]
                                                lista_61 = (lista_6['src'])
                                                lista_62 = lista_61.replace('_US40_', '_SX679_')
                                                lista_62 = lista_62.replace('_SX38_SY50_CR,0,0,38,50_', '_AC_SX679_')
                                                lista_62 = lista_62.replace('_AC_SR38,50_', '_AC_SX679_')
                                                ws.cell(column=(30), row=newRowLocation, value=lista_62)
                                                print(lista_62)
                                                if lista[7] != None:
                                                    lista_7 = lista[7]
                                                    lista_71 = (lista_7['src'])
                                                    lista_72 = lista_71.replace('_US40_', '_SX679_')
                                                    lista_72 = lista_72.replace('_SX38_SY50_CR,0,0,38,50_',
                                                                                '_AC_SX679_')
                                                    lista_72 = lista_72.replace('_AC_SR38,50_', '_AC_SX679_')
                                                    ws.cell(column=(31), row=newRowLocation, value=lista_72)
                                                    print(lista_72)
                                                    if lista[8] != None:
                                                        lista_8 = lista[8]
                                                        lista_81 = (lista_8['src'])
                                                        lista_82 = lista_81.replace('_US40_', '_SX679_')
                                                        lista_82 = lista_82.replace('_SX38_SY50_CR,0,0,38,50_',
                                                                                    '_AC_SX679_')
                                                        lista_82 = lista_82.replace('_AC_SR38,50_', '_AC_SX679_')
                                                        ws.cell(column=(32), row=newRowLocation, value=lista_82)
                                                        print(lista_82)
                                                        if lista[9] != None:
                                                            lista_9 = lista[9]
                                                            lista_91 = (lista_9['src'])
                                                            lista_92 = lista_91.replace('_US40_', '_SX679_')
                                                            lista_92 = lista_92.replace('_SX38_SY50_CR,0,0,38,50_',
                                                                                        '_AC_SX679_')
                                                            lista_92 = lista_92.replace('_AC_SR38,50_', '_AC_SX679_')
                                                            ws.cell(column=(33), row=newRowLocation, value=lista_92)
                                                            print(lista_92)

                    ws.cell(column=3, row=newRowLocation, value=categoria)
                    ws.cell(column=4, row=newRowLocation, value=title)
                    ws.cell(column=5, row=newRowLocation, value=desc_4)
                    ws.cell(column=6, row=newRowLocation, value=int(converted_price))
                    ws.cell(column=7, row=newRowLocation, value=int(formula))
                    ws.cell(column=8, row=newRowLocation, value=amazin)
                    ws.cell(column=9, row=newRowLocation, value=estado)
                    ws.cell(column=10, row=newRowLocation, value=int(stock))
                    ws.cell(column=11, row=newRowLocation, value=disponibilidad)
                    ws.cell(column=12, row=newRowLocation, value=tipo)
                    ws.cell(column=13, row=newRowLocation, value=condicion)
                    ws.cell(column=14, row=newRowLocation, value=envio)
                    ws.cell(column=16, row=newRowLocation, value=modo)
                    ws.cell(column=17, row=newRowLocation, value=metodo)
                    ws.cell(column=18, row=newRowLocation, value=retiro)
                    ws.cell(column=19, row=newRowLocation, value=garantia)
                    #		ws.cell(column=24,row=newRowLocation, value=amgaz[0])
                    #		ws.cell(column=25,row=newRowLocation, value=amgaz[1])
                    #		ws.cell(column=26,row=newRowLocation, value=amgaz[2])
                    #		ws.cell(column=27,row=newRowLocation, value=amgaz[3])
                    #		ws.cell(column=28,row=newRowLocation, value=amgaz[4])
                    #		ws.cell(column=29,row=newRowLocation, value=amgaz[5])
                    #		ws.cell(column=30,row=newRowLocation, value=amgaz[6])
                    #		ws.cell(column=31,row=newRowLocation, value=amgaz[7])
                    #		ws.cell(column=32,row=newRowLocation, value=amgaz[8])
                    ws.cell(column=43, row=newRowLocation, value=marca)
                    ws.cell(column=44, row=newRowLocation, value=amazin)

                    wb.save('test.xlsx')

                else:
                    print('precio fuera de rango')
            else:
                print('No es enviado por Amazon')

