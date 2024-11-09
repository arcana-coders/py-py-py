import openpyxl
import requests
from bs4 import BeautifulSoup as soup
import time
from amazoncaptcha import AmazonCaptcha

# from colorama import Fore, init


def precioydisponibilidad():
    # se extrae el precio y se le da formato, se quita la coma, y las decimales y el simbolo
    price_1 = scrap.find(id='priceblock_ourprice')
    if price_1 is None:
        price_1 = scrap.find(id='priceblock_saleprice')
    if price_1 is None:
        price_1 = scrap.find(id='priceblock_dealprice')
    if price_1 is None:
        price_1 = scrap.find(id='newBuyBoxPrice')
    if price_1 is None:
        ws.cell(row, 8).value = 0
        return
    # print(price_1)
    price_4 = price_1.get_text()
    verted_price = price_4[1:10]
    onverted_price = verted_price.replace(',', '')
    sep = '.'
    converted_price = onverted_price.split(sep, 1)[0]
    ws.cell(row, 5).value = int(converted_price)
    print('precio proveedor: ' + str(converted_price))

    # arreglamos el precio con la formula
    rmula = (int(converted_price) * float(1.35))
    ormula = str(rmula)
    formula = ormula.split(sep, 1)[0]
    ws.cell(row, 6).value = int(formula)

    print('precio de venta: ' + formula)

    #La cantidad minima de compra cuando exista.
    cantminima = scrap.find(id='quantity')
    if cantminima is None:
        ws.cell(row, 8).value = 1
        print('Hay solo 1 pieza disponible')
    elif cantminima is not None:
        cantminima = cantminima.get_text().split()
        # print(cantminima)
        if cantminima[0] == 'Selecciona':
            print('Cantidad minima de compra es: ' + cantminima[3])
            formula = int(formula) * int(cantminima[3])
            ws.cell(row, 6).value = int(formula)
            print('Nuevo precio es: ' + str(formula) + ' por ' + str(cantminima[3]) + ' piezas.')

            stock = 1
            ws.cell(row, 8).value = int(stock)
            print(cantminima)
            print('Hay ' + str(stock) + ' piezas disponibles.')
        if cantminima[0] == '1':
            stock = cantminima[-1]
            print('Hay ' + str(stock) + ' piezas disponibles.')
            ws.cell(row, 8).value = int(stock)


    # Ahora la disponibilidad cuando es nacional o importado.
    disponibilidad = scrap.find(id='priceblock_ourprice_ifdmsg')
    if disponibilidad != None:
        disponibilidad = '7 a 10 dias. Producto de importacion'
    if disponibilidad == None:
        disponibilidad = '2 a 5 dias. Envio Nacional'
    ws.cell(row, 9).value = disponibilidad

    print('Disponible en: ' + disponibilidad + ' dias.')

# Primero las variables.

wb = openpyxl.load_workbook("woo-yasubidos.xlsx")


# Comenzamos el loop donde lee las columnas desde la celda 2 hasta donde tenga valor
for sheet in wb.worksheets:
    ws = sheet
    print('')
    print('')
    print(sheet)
    print('')
    print('')
    for row in range(2, ws.max_row):
        if (ws.cell(row, 7).value is None):
            break

        Asin = (ws.cell(row, 7).value)
        # estado = (ws.cell(row, 9).value)  # sus opciones son 'Pausada' o 'Activa'
        pre_pro = (ws.cell(row, 5).value)
        precio = (ws.cell(row, 6).value)
        # estate = {"1": "Activa", "2": "Pausada"}
        # activa = ('Activa')
        # pausada = 'Pausada'
        time.sleep(3)
        URL = 'https://www.amazon.com.mx/Morama-Harina-Para-Hotcakes-Gramos/dp/' + Asin + '/ref=pd_bxgy_img_2/136-5060219-4616706?_encoding=UTF8&pd_rd_i=B07ZHL7HNF&pd_rd_r=801c462e-011d-4d08-b788-6cee9b966be0&pd_rd_w=rCaD9&pd_rd_wg=9BLh6&pf_rd_p=9c27df7e-1c11-4701-82a3-a796c9270e4c&pf_rd_r=3BABYST5G5BTKF4G2026&psc=1&refRID=3BABYST5G5BTKF4G2026'
        headers = {
            "User-Agent": 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.190 Safari/537.36'}

        print('')
        print(URL)
        # print('Estado anterior: ' + estado)
        page = requests.get(URL, headers=headers)
        print('conexion: ' + str(page.status_code))
        scrap = soup(page.content, 'html.parser')

        #print(scrap.text)

        disponible = scrap.find(id='availability')
        enviado = scrap.find(id='merchant-info')
        enviado1 = scrap.find(id="buybox-tabular")
        # print(enviado)
        # init(autoreset=True)
        # dispo = 'Activa'
        # nodispo = 'Pausada'

        if disponible == None:
            print('No tiene informacion.')
            ws.cell(row, 8).value = 0
            print('no disponible')
        elif disponible is not None:
            disponible = disponible.get_text().split()
            # print(disponible)
            try:
                if disponible[0] == ('No'):
                    ws.cell(row, 8).value = 0
                    print('No disponible')
                    print('Motivo: ' + str(disponible))
                elif disponible[0] == ('Disponible'):
                    ws.cell(row, 8).value = 0
                    print('No disponible')
                    print('Motivo: ' + str(disponible))
                elif disponible[0] == ('Disponible.'):
                    ws.cell(row, 8).value = 3
                    print('Disponible')
                    print('Motivo: ' + str(disponible))

                    if enviado == None and enviado1 == None:

                        ws.cell(row, 8).value = 0
                        print('No disponible')
                        print('Motivo: ' + str(enviado))

                    elif enviado is not None:
                        enviado = enviado.get_text().split()

                        if any(ele == 'Amazon' for ele in enviado) or any(ele == 'Amazon.' for ele in enviado):
                            ws.cell(row, 8).value = 3
                            enviado = " ".join(enviado)
                            print('Disponible')
                            print('Motivo: ' + str(enviado))

                            precioydisponibilidad()
                        else:
                            ws.cell(row, 8).value = 0
                            enviado = " ".join(enviado)
                            print('No disponible')
                            print('Motivo: ' + str(enviado))

                        disponible = " ".join(disponible)
                        # print(disponible)
                    elif enviado1 is not None:
                        enviado1 = enviado1.get_text().split()

                        if any(ele == 'Amazon' for ele in enviado1) or any(ele == 'Amazon.' for ele in enviado1):
                            ws.cell(row, 8).value = 3
                            enviado1 = " ".join(enviado1)
                            print('Disponible')
                            print('Motivo: ' + str(enviado1))

                            precioydisponibilidad()
                        else:
                            ws.cell(row, 8).value = 0
                            enviado1 = " ".join(enviado1)
                            print('No disponible')
                            print('Motivo: ' + str(enviado1))

                        disponible = " ".join(disponible)
                        # print(disponible)

                else:
                    ws.cell(row, 8).value = 0
                    print('No disponible')
                    print('Motivo: ' + str(disponible))

            except Exception as e:
                print('Error... Error.... En linea 90 seguramente.')
                pass

    wb.save('woo-yasubidos.xlsx')
    print('')
    print('')
    print('Guardando.... guardando ....')
    print('')
    print('')
